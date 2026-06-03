"""
MathTestUZ API Views
All views for auth, teacher panel, student panel, and AI chatbot.
"""

import random
import os
import io
from datetime import datetime

from django.utils import timezone
from django.db.models import Count, Avg, Q, Sum
from django.http import HttpResponse
from rest_framework import status, generics, permissions, serializers
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework_simplejwt.tokens import RefreshToken

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .models import (User, Test, Question, Attempt, Answer,
                     Material, ChatSession, ChatMessage, MatchingPair,
                     MaterialReadLog, Classroom)
from .serializers import (
    RegisterSerializer, UserSerializer, ProfileUpdateSerializer, TestSerializer,
    QuestionSerializer, QuestionStudentSerializer,
    AttemptSerializer, AnswerSerializer, MaterialSerializer,
    ChatMessageSerializer, ChatSessionSerializer, MatchingPairSerializer,
    MaterialReadLogSerializer, ClassroomSerializer
)
from .permissions import IsTeacher, IsStudent


# ═══════════════════════════════════════════════════════════════════════════════
# AUTH VIEWS
# ═══════════════════════════════════════════════════════════════════════════════

class RegisterView(generics.CreateAPIView):
    """POST /api/auth/register/ — Create new user account."""
    queryset = User.objects.all()
    serializer_class = RegisterSerializer
    permission_classes = [permissions.AllowAny]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        # Return JWT tokens immediately after registration
        refresh = RefreshToken.for_user(user)
        return Response({
            'user': UserSerializer(user).data,
            'access': str(refresh.access_token),
            'refresh': str(refresh),
        }, status=status.HTTP_201_CREATED)


class LoginView(APIView):
    """POST /api/auth/login/ — Login and get JWT tokens."""
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        from django.contrib.auth import authenticate
        username = request.data.get('username')
        password = request.data.get('password')
        user = authenticate(username=username, password=password)
        if not user:
            return Response({'detail': 'Username yoki parol noto\'g\'ri!'},
                            status=status.HTTP_401_UNAUTHORIZED)
        refresh = RefreshToken.for_user(user)
        return Response({
            'user': UserSerializer(user).data,
            'access': str(refresh.access_token),
            'refresh': str(refresh),
        })


class MeView(APIView):
    """GET /api/auth/me/ — Get current user info.
       PATCH /api/auth/me/ — Update profile (full_name, class_name, password).
    """

    def get(self, request):
        return Response(UserSerializer(request.user).data)

    def patch(self, request):
        serializer = ProfileUpdateSerializer(
            request.user, data=request.data, partial=True
        )
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        return Response(UserSerializer(user).data)


# ═══════════════════════════════════════════════════════════════════════════════
# TEACHER VIEWS
# ═══════════════════════════════════════════════════════════════════════════════

class TeacherDashboardView(APIView):
    """GET /api/teacher/dashboard/ — Stats for teacher."""
    permission_classes = [IsTeacher]

    def get(self, request):
        teacher = request.user
        tests = Test.objects.filter(teacher=teacher)
        total_tests = tests.count()
        total_students = User.objects.filter(role='student').count()
        total_attempts = Attempt.objects.filter(test__teacher=teacher, is_completed=True).count()

        # Average score per test
        test_stats = []
        for test in tests:
            attempts = Attempt.objects.filter(test=test, is_completed=True)
            totals = attempts.aggregate(s=Sum('score'), t=Sum('total_questions'))
            s = totals['s'] or 0
            t = totals['t'] or 0
            avg_pct = round((s / t * 100) if t > 0 else 0, 1)
            test_stats.append({
                'test_id': test.id,
                'test_name': test.name,
                'attempts_count': attempts.count(),
                'avg_percentage': avg_pct,
            })

        # Hardest questions (most wrong answers)
        hard_questions = (
            Answer.objects.filter(
                attempt__test__teacher=teacher,
                is_correct=False
            )
            .values('question__id', 'question__text', 'question__test__name')
            .annotate(wrong_count=Count('id'))
            .order_by('-wrong_count')[:5]
        )

        # Class-wise stats
        class_stats = (
            Attempt.objects.filter(test__teacher=teacher, is_completed=True)
            .values('student__class_name')
            .annotate(
                count=Count('id'),
                avg_score=Avg('score')
            )
            .order_by('student__class_name')
        )

        return Response({
            'total_tests': total_tests,
            'total_students': total_students,
            'total_attempts': total_attempts,
            'test_stats': test_stats,
            'hard_questions': list(hard_questions),
            'class_stats': list(class_stats),
        })


# ─── Classrooms (Teacher) ──────────────────────────────────────────────────────

import string

def generate_invite_code():
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))

class ClassroomListCreateView(generics.ListCreateAPIView):
    """GET/POST /api/teacher/classrooms/"""
    serializer_class = ClassroomSerializer
    permission_classes = [IsTeacher]

    def get_queryset(self):
        return Classroom.objects.filter(teacher=self.request.user).order_by('-created_at')

    def perform_create(self, serializer):
        code = generate_invite_code()
        while Classroom.objects.filter(invite_code=code).exists():
            code = generate_invite_code()
        serializer.save(teacher=self.request.user, invite_code=code)

class ClassroomDetailView(generics.RetrieveUpdateDestroyAPIView):
    """GET/PUT/PATCH/DELETE /api/teacher/classrooms/<id>/"""
    serializer_class = ClassroomSerializer
    permission_classes = [IsTeacher]

    def get_queryset(self):
        return Classroom.objects.filter(teacher=self.request.user)

class ClassroomStudentsView(generics.ListAPIView):
    """GET /api/teacher/classrooms/<id>/students/"""
    serializer_class = UserSerializer
    permission_classes = [IsTeacher]

    def get_queryset(self):
        classroom = Classroom.objects.filter(id=self.kwargs['pk'], teacher=self.request.user).first()
        if not classroom:
            return User.objects.none()
        return classroom.students.order_by('full_name', 'username')

class AddStudentToClassroomView(APIView):
    """POST /api/teacher/classrooms/<id>/add-student/"""
    permission_classes = [IsTeacher]

    def post(self, request, pk):
        classroom = Classroom.objects.filter(id=pk, teacher=request.user).first()
        if not classroom:
            return Response({'detail': 'Sinf topilmadi'}, status=404)
        
        student_id = request.data.get('student_id')
        username = request.data.get('username')
        
        if student_id:
            student = User.objects.filter(id=student_id, role='student').first()
        elif username:
            student = User.objects.filter(username=username, role='student').first()
        else:
            return Response({'detail': 'student_id yoki username berilishi shart'}, status=400)
            
        if not student:
            return Response({'detail': 'O\'quvchi topilmadi'}, status=404)
            
        classroom.students.add(student)
        return Response({'detail': 'Qo\'shildi'})

class RemoveStudentFromClassroomView(APIView):
    """POST /api/teacher/classrooms/<id>/remove-student/"""
    permission_classes = [IsTeacher]

    def post(self, request, pk):
        classroom = Classroom.objects.filter(id=pk, teacher=request.user).first()
        if not classroom:
            return Response({'detail': 'Sinf topilmadi'}, status=404)
            
        student_id = request.data.get('student_id')
        classroom.students.remove(student_id)
        return Response({'detail': 'O\'chirildi'})

class AllStudentsView(generics.ListAPIView):
    """GET /api/teacher/students/"""
    serializer_class = UserSerializer
    permission_classes = [IsTeacher]
    def get_queryset(self):
        return User.objects.filter(role='student').order_by('full_name', 'username')

class TeacherClassroomLeaderboardView(APIView):
    """GET /api/teacher/classrooms/<id>/leaderboard/"""
    permission_classes = [IsTeacher]

    def get(self, request, pk):
        classroom = Classroom.objects.filter(id=pk, teacher=request.user).first()
        if not classroom:
            return Response({'detail': 'Sinf topilmadi'}, status=404)
        
        students = classroom.students.annotate(
            classroom_score=Sum('attempts__score', filter=Q(attempts__test__classroom=classroom, attempts__is_completed=True))
        )
        
        board = []
        for s in students:
            score = s.classroom_score or 0
            board.append({
                'student_id': s.id,
                'full_name': s.full_name or s.username,
                'classroom_xp': score * 10,
                'total_xp': s.total_xp,
                'level': s.level
            })
            
        board.sort(key=lambda x: -x['classroom_xp'])
        for i, row in enumerate(board):
            row['rank'] = i + 1
            
        return Response(board)


# ─── Tests ─────────────────────────────────────────────────────────────────────

class TeacherTestListCreateView(generics.ListCreateAPIView):
    """GET/POST /api/teacher/tests/"""
    serializer_class = TestSerializer
    permission_classes = [IsTeacher]

    def get_queryset(self):
        return Test.objects.filter(teacher=self.request.user).order_by('-created_at')

    def perform_create(self, serializer):
        serializer.save(teacher=self.request.user)


class TeacherTestDetailView(generics.RetrieveUpdateDestroyAPIView):
    """GET/PUT/PATCH/DELETE /api/teacher/tests/<id>/"""
    serializer_class = TestSerializer
    permission_classes = [IsTeacher]

    def get_queryset(self):
        return Test.objects.filter(teacher=self.request.user)


class TeacherTestResultsView(APIView):
    """GET /api/teacher/tests/<id>/results/ — All attempts for a test."""
    permission_classes = [IsTeacher]

    def get(self, request, pk):
        test = Test.objects.filter(pk=pk, teacher=request.user).first()
        if not test:
            return Response({'detail': 'Test topilmadi.'}, status=404)
        attempts = Attempt.objects.filter(test=test, is_completed=True).select_related('student')
        data = []
        for a in attempts:
            data.append({
                'attempt_id': a.id,
                'student': a.student.full_name or a.student.username,
                'class_name': a.student.class_name,
                'score': a.score,
                'total': a.total_questions,
                'percentage': a.percentage,
                'grade': a.grade,
                'submitted_at': a.submitted_at,
            })
        return Response(data)


class ExportResultsExcelView(APIView):
    """GET /api/teacher/tests/<id>/export/ — Download XLSX of results."""
    permission_classes = [IsTeacher]

    def get(self, request, pk):
        test = Test.objects.filter(pk=pk, teacher=request.user).first()
        if not test:
            return Response({'detail': 'Test topilmadi.'}, status=404)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Natijalar"

        # Header style
        header_fill = PatternFill("solid", fgColor="1E40AF")
        header_font = Font(color="FFFFFF", bold=True)

        headers = ["O'quvchi", "Sinf", "Ball", "Jami", "Foiz (%)", "Baho", "Material o'qilganmi?", "Topshirilgan"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Pre-fetch material read logs for this test's material
        material = test.material
        read_student_ids = set()
        if material:
            read_student_ids = set(
                MaterialReadLog.objects.filter(material=material)
                .values_list('student_id', flat=True)
            )

        attempts = Attempt.objects.filter(test=test, is_completed=True).select_related('student')
        for row, a in enumerate(attempts, 2):
            read_status = '✅ Ha' if a.student_id in read_student_ids else '❌ Yo\'q'
            ws.cell(row=row, column=1, value=a.student.full_name or a.student.username)
            ws.cell(row=row, column=2, value=a.student.class_name or '-')
            ws.cell(row=row, column=3, value=a.score)
            ws.cell(row=row, column=4, value=a.total_questions)
            ws.cell(row=row, column=5, value=a.percentage)
            ws.cell(row=row, column=6, value=a.grade)
            ws.cell(row=row, column=7, value=read_status)
            ws.cell(row=row, column=8, value=str(a.submitted_at)[:19] if a.submitted_at else '-')

        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{test.name}_natijalar.xlsx"'
        wb.save(response)
        return response


# ─── Questions ─────────────────────────────────────────────────────────────────

class TeacherQuestionListCreateView(generics.ListCreateAPIView):
    """GET/POST /api/teacher/tests/<test_id>/questions/"""
    serializer_class = QuestionSerializer
    permission_classes = [IsTeacher]

    def get_queryset(self):
        return Question.objects.filter(
            test__teacher=self.request.user,
            test_id=self.kwargs['test_id']
        )

    def perform_create(self, serializer):
        test = Test.objects.get(pk=self.kwargs['test_id'], teacher=self.request.user)
        
        # Enforce that a test can only have ONE type of question
        # i.e., multiple_choice only OR matching_pairs only
        new_q_type = serializer.validated_data.get('question_type', 'multiple_choice')
        existing_question = test.questions.first()
        
        if existing_question and existing_question.question_type != new_q_type:
            raise serializers.ValidationError({
                "question_type": f"Bu testda allaqachon '{existing_question.get_question_type_display()}' turidagi savollar mavjud. "
                                 f"Boshqa turdagi savol qo'shish mumkin emas."
            })

        serializer.save(test=test)


class TeacherQuestionDetailView(generics.RetrieveUpdateDestroyAPIView):
    """GET/PUT/PATCH/DELETE /api/teacher/questions/<id>/"""
    serializer_class = QuestionSerializer
    permission_classes = [IsTeacher]

    def get_queryset(self):
        return Question.objects.filter(test__teacher=self.request.user)


class MatchingPairListCreateView(generics.ListCreateAPIView):
    """GET/POST /api/teacher/questions/<question_id>/matching-pairs/"""
    serializer_class = MatchingPairSerializer
    permission_classes = [IsTeacher]

    def get_queryset(self):
        question = Question.objects.filter(
            test__teacher=self.request.user,
            id=self.kwargs['question_id']
        ).first()
        if not question:
            return MatchingPair.objects.none()
        return MatchingPair.objects.filter(question=question)

    def perform_create(self, serializer):
        question = Question.objects.get(
            pk=self.kwargs['question_id'],
            test__teacher=self.request.user
        )
        serializer.save(question=question)


class MatchingPairDetailView(generics.RetrieveUpdateDestroyAPIView):
    """GET/PUT/PATCH/DELETE /api/teacher/matching-pairs/<id>/"""
    serializer_class = MatchingPairSerializer
    permission_classes = [IsTeacher]

    def get_queryset(self):
        return MatchingPair.objects.filter(
            question__test__teacher=self.request.user
        )


# ─── Materials ─────────────────────────────────────────────────────────────────

class TeacherMaterialListCreateView(generics.ListCreateAPIView):
    """GET/POST /api/teacher/materials/"""
    serializer_class = MaterialSerializer
    permission_classes = [IsTeacher]

    def get_queryset(self):
        return Material.objects.filter(teacher=self.request.user).order_by('-created_at')

    def perform_create(self, serializer):
        serializer.save(teacher=self.request.user)


class TeacherMaterialDetailView(generics.RetrieveUpdateDestroyAPIView):
    """GET/PUT/PATCH/DELETE /api/teacher/materials/<id>/"""
    serializer_class = MaterialSerializer
    permission_classes = [IsTeacher]

    def get_queryset(self):
        return Material.objects.filter(teacher=self.request.user)


# ═══════════════════════════════════════════════════════════════════════════════
# STUDENT VIEWS
# ═══════════════════════════════════════════════════════════════════════════════

# ─── Classrooms (Student) ──────────────────────────────────────────────────────

class JoinClassroomView(APIView):
    """POST /api/student/join-classroom/"""
    permission_classes = [IsStudent]

    def post(self, request):
        code = request.data.get('invite_code')
        if not code:
            return Response({'detail': 'Taklif kodi kiritilmagan'}, status=400)
            
        classroom = Classroom.objects.filter(invite_code=code).first()
        if not classroom:
            return Response({'detail': 'Sinf topilmadi (kod noto\'g\'ri)'}, status=404)
            
        if classroom.students.filter(id=request.user.id).exists():
            return Response({'detail': 'Siz allaqachon bu sinfga qo\'shilgansiz'}, status=400)
            
        classroom.students.add(request.user)
        return Response({
            'detail': f"'{classroom.name}' sinfiga muvaffaqiyatli qo'shildingiz",
            'classroom': ClassroomSerializer(classroom).data
        })

class MyClassroomsView(generics.ListAPIView):
    """GET /api/student/my-classrooms/"""
    serializer_class = ClassroomSerializer
    permission_classes = [IsStudent]

    def get_queryset(self):
        return self.request.user.classrooms.all().order_by('-created_at')

class StudentMaterialsView(generics.ListAPIView):
    """GET /api/student/materials/ — Materials available to student's class or public."""
    serializer_class = MaterialSerializer
    permission_classes = [permissions.AllowAny]

    def get_queryset(self):
        user = self.request.user
        if user.is_authenticated:
            return Material.objects.filter(
                Q(class_name=user.class_name, class_name__isnull=False) |
                Q(classroom__in=user.classrooms.all()) |
                Q(access_type='public')
            ).distinct().order_by('-created_at')
        return Material.objects.filter(access_type='public').order_by('-created_at')


class StudentMarkMaterialReadView(APIView):
    """POST /api/student/materials/<material_id>/read/ — Mark material as read."""
    permission_classes = [permissions.AllowAny]

    def post(self, request, material_id):
        material = Material.objects.filter(pk=material_id).first()
        if not material:
            return Response({'detail': 'Material topilmadi.'}, status=404)
        # For authenticated students, save the log. For guests, just return success.
        if request.user.is_authenticated:
            # get_or_create prevents duplicates
            log, created = MaterialReadLog.objects.get_or_create(
                student=request.user,
                material=material
            )
            read_at = log.read_at
            already_read = not created
        else:
            read_at = timezone.now()
            already_read = False

        return Response({
            'material_id': material.id,
            'read_at': read_at,
            'already_read': already_read,
            # return linked test id if any, so frontend can navigate
            'linked_test_id': material.tests.filter(is_active=True).values_list('id', flat=True).first()
        })


class StudentTestListView(generics.ListAPIView):
    """GET /api/student/tests/ — Tests available to this student or public."""
    serializer_class = TestSerializer
    permission_classes = [permissions.AllowAny]

    def get_queryset(self):
        user = self.request.user
        if user.is_authenticated:
            return Test.objects.filter(
                is_active=True
            ).filter(
                Q(access_type='public') | 
                Q(allowed_class=user.class_name) |
                Q(classroom__in=user.classrooms.all())
            ).distinct().order_by('-created_at')
        return Test.objects.filter(is_active=True, access_type='public').order_by('-created_at')


class StudentStartAttemptView(APIView):
    """POST /api/student/tests/<test_id>/start/ — Start a new attempt."""
    permission_classes = [IsStudent]

    def post(self, request, test_id):
        student = request.user
        test = Test.objects.filter(pk=test_id, is_active=True).first()
        if not test:
            return Response({'detail': 'Test topilmadi.'}, status=404)

        # Check access
        if test.access_type == 'class':
            has_access = False
            if test.allowed_class and student.class_name == test.allowed_class:
                has_access = True
            elif test.classroom and student.classrooms.filter(id=test.classroom.id).exists():
                has_access = True
                
            if not has_access:
                return Response({'detail': 'Bu testga ruxsat yo\'q.'}, status=403)

        # Check attempt count
        attempts_count = Attempt.objects.filter(
            student=student, test=test, is_completed=True).count()
        if test.max_attempts and attempts_count >= test.max_attempts:
            return Response({'detail': f'Maksimal urinishlar soni ({test.max_attempts}) tugadi.'},
                            status=400)

        # Check for existing incomplete attempt
        existing = Attempt.objects.filter(student=student, test=test, is_completed=False).first()
        if existing:
            questions = Question.objects.filter(id__in=existing.question_order)
            q_map = {q.id: q for q in questions}
            ordered_qs = [q_map[qid] for qid in existing.question_order if qid in q_map]
            return Response({
                'attempt_id': existing.id,
                'questions': QuestionStudentSerializer(ordered_qs, many=True).data,
                'time_limit': test.time_limit,
                'chatbot_mode': test.chatbot_mode,
                'started_at': existing.started_at,
            })

        # Create new attempt with randomized questions
        question_ids = list(Question.objects.filter(test=test).values_list('id', flat=True))
        random.shuffle(question_ids)

        questions = Question.objects.filter(id__in=question_ids).prefetch_related('matching_pairs')
        q_map = {q.id: q for q in questions}
        ordered_qs = [q_map[qid] for qid in question_ids if qid in q_map]

        # Total points = sum of pairs for matching_pairs, 1 for multiple_choice
        total_points = sum(
            q.matching_pairs.count() if q.question_type == 'matching_pairs' else 1
            for q in ordered_qs
        )

        attempt = Attempt.objects.create(
            student=student,
            test=test,
            total_questions=total_points,
            question_order=question_ids,
        )

        return Response({
            'attempt_id': attempt.id,
            'questions': QuestionStudentSerializer(ordered_qs, many=True).data,
            'time_limit': test.time_limit,
            'chatbot_mode': test.chatbot_mode,
            'started_at': attempt.started_at,
        }, status=201)


class StudentSubmitAnswerView(APIView):
    """POST /api/student/attempts/<attempt_id>/answer/ — Submit one answer."""
    permission_classes = [IsStudent]

    def post(self, request, attempt_id):
        attempt = Attempt.objects.filter(
            pk=attempt_id, student=request.user, is_completed=False).first()
        if not attempt:
            return Response({'detail': 'Urinish topilmadi yoki tugallangan.'}, status=404)

        question_id = request.data.get('question_id')
        # for matching pairs we expect a mapping in selected_matching
        selected_matching = request.data.get('selected_matching')
        selected_option = request.data.get('selected_option', '')
        if isinstance(selected_option, str):
            selected_option = selected_option.upper()

        question = Question.objects.filter(pk=question_id, test=attempt.test).first()
        if not question:
            return Response({'detail': 'Savol topilmadi.'}, status=404)

        # Create or update the answer
        defaults = {}
        if question.question_type == 'matching_pairs':
            defaults['selected_matching'] = selected_matching
        else:
            defaults['selected_option'] = selected_option

        answer, created = Answer.objects.get_or_create(
            attempt=attempt, question=question,
            defaults=defaults
        )
        if not created:
            if question.question_type == 'matching_pairs':
                answer.selected_matching = selected_matching
            else:
                answer.selected_option = selected_option
            answer.save()

        resp = {'is_correct': answer.is_correct,
                'explanation': question.explanation if not answer.is_correct else ''}
        if question.question_type != 'matching_pairs':
            resp['correct_option'] = question.correct_option
        return Response(resp)


class StudentFinishAttemptView(APIView):
    """POST /api/student/attempts/<attempt_id>/finish/ — Complete the attempt."""
    permission_classes = [IsStudent]

    def post(self, request, attempt_id):
        attempt = Attempt.objects.filter(
            pk=attempt_id, student=request.user, is_completed=False).first()
        if not attempt:
            return Response({'detail': 'Urinish topilmadi yoki tugallangan.'}, status=404)

        # Per-pair scoring: matching_pairs questions give 1 point per correct pair
        answers = Answer.objects.filter(attempt=attempt).select_related('question').prefetch_related('question__matching_pairs')
        earned_points = 0
        for ans in answers:
            q = ans.question
            if q.question_type == 'matching_pairs':
                if ans.selected_matching:
                    pairs = list(q.matching_pairs.all())
                    for p in pairs:
                        student_val = (ans.selected_matching.get(str(p.id))
                                       or ans.selected_matching.get(p.id))
                        if student_val is not None and str(student_val) == str(p.id):
                            earned_points += 1
            else:
                if ans.is_correct:
                    earned_points += 1

        attempt.score = earned_points
        # Do not recalculate attempt.total_questions — it was correctly set when the attempt was started.
        attempt.is_completed = True
        attempt.submitted_at = timezone.now()
        attempt.save()

        # Gamification: Update user XP and Level
        student = request.user
        xp_earned = earned_points * 10
        student.total_xp += xp_earned
        student.level = (student.total_xp // 100) + 1
        student.save()

        # Detailed results: each question + student's answer
        answers = Answer.objects.filter(attempt=attempt).select_related('question')
        wrong_details = []
        for ans in answers:
            if not ans.is_correct:
                if ans.question.question_type == 'matching_pairs':
                    wrong_details.append({
                        'question': ans.question.text,
                        'your_answer': ans.selected_matching,
                        'correct_answer': 'Har bir element o\'z juftiga mos kelishi kerak',
                        'explanation': ans.question.explanation,
                    })
                else:
                    wrong_details.append({
                        'question': ans.question.text,
                        'your_answer': ans.selected_option,
                        'correct_answer': ans.question.correct_option,
                        'explanation': ans.question.explanation,
                    })

        return Response({
            'score': attempt.score,
            'total': attempt.total_questions,
            'percentage': attempt.percentage,
            'grade': attempt.grade,
            'wrong_details': wrong_details,
            'xp_earned': xp_earned,
            'new_level': student.level,
            'total_xp': student.total_xp,
        })


class StudentHistoryView(APIView):
    """GET /api/student/history/ — Student's attempt history."""
    permission_classes = [IsStudent]

    def get(self, request):
        attempts = Attempt.objects.filter(
            student=request.user, is_completed=True
        ).select_related('test').order_by('-submitted_at')

        data = []
        for a in attempts:
            data.append({
                'attempt_id': a.id,
                'test_name': a.test.name,
                'score': a.score,
                'total': a.total_questions,
                'percentage': a.percentage,
                'grade': a.grade,
                'submitted_at': a.submitted_at,
            })
        return Response(data)


class StudentLeaderboardView(APIView):
    """GET /api/student/leaderboard/ — Top students by avg score or global XP."""
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        if not request.user.is_authenticated:
            # Global leaderboard for guests
            students = User.objects.filter(role='student').order_by('-total_xp')[:10]
            board = []
            for i, s in enumerate(students):
                board.append({
                    'rank': i + 1,
                    'full_name': s.full_name or s.username,
                    'total_xp': s.total_xp,
                    'level': s.level,
                    'class_name': s.class_name or '-'
                })
            return Response(board)
            
        classrooms = request.user.classrooms.all()
        class_name = request.user.class_name
        
        if classrooms.exists():
            students = User.objects.filter(role='student', classrooms__in=classrooms).distinct()
        elif class_name:
            students = User.objects.filter(role='student', class_name=class_name)
        else:
            students = User.objects.filter(id=request.user.id) # Faqat o'zi
            
        board = []
        for s in students:
            attempts = Attempt.objects.filter(student=s, is_completed=True)
            count = attempts.count()
            if count == 0:
                continue
            avg_pct = round(sum(a.percentage for a in attempts) / count, 1)
            total_score = sum(a.score for a in attempts)
            board.append({
                'student_id': s.id,
                'full_name': s.full_name or s.username,
                'attempts_count': count,
                'avg_percentage': avg_pct,
                'total_score': total_score,
                'is_me': s.id == request.user.id,
            })
        board.sort(key=lambda x: (-x['avg_percentage'], -x['total_score']))
        for i, row in enumerate(board):
            row['rank'] = i + 1
        return Response(board)


class StudentProgressView(APIView):
    """GET /api/student/progress/ — Material & test progress for student."""
    permission_classes = [IsStudent]

    def get(self, request):
        student = request.user
        if not student.is_authenticated:
            return Response(status=401)
        class_name = student.class_name

        total_materials = Material.objects.filter(
            Q(class_name=class_name) | Q(class_name='') | Q(classroom__in=student.classrooms.all())
        ).distinct().count()
        read_materials = MaterialReadLog.objects.filter(student=student).count()

        total_tests = Test.objects.filter(
            is_active=True
        ).filter(
            Q(access_type='public') | Q(allowed_class=class_name) | Q(classroom__in=student.classrooms.all())
        ).distinct().count()
        completed_tests = Attempt.objects.filter(
            student=student, is_completed=True
        ).values('test').distinct().count()

        # Last 7 attempts for mini trend
        recent = Attempt.objects.filter(
            student=student, is_completed=True
        ).order_by('-submitted_at')[:7]
        trend = [{'test': a.test.name, 'pct': a.percentage, 'date': str(a.submitted_at)[:10]} for a in reversed(list(recent))]

        return Response({
            'total_materials': total_materials,
            'read_materials': read_materials,
            'total_tests': total_tests,
            'completed_tests': completed_tests,
            'trend': trend,
        })


class ImportQuestionsFromExcelView(APIView):
    """POST /api/teacher/tests/<pk>/import-questions/ — Bulk import from .xlsx"""
    permission_classes = [IsTeacher]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, pk):
        test = Test.objects.filter(pk=pk, teacher=request.user).first()
        if not test:
            return Response({'detail': 'Test topilmadi.'}, status=404)

        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'detail': 'Excel fayl yuklanishi shart.'}, status=400)

        try:
            wb = openpyxl.load_workbook(file_obj)
            ws = wb.active
        except Exception:
            return Response({'detail': 'Excel faylni o\'qib bo\'lmadi.'}, status=400)

        created = 0
        errors = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Expected columns: text | A | B | C | D | correct | explanation | difficulty
            if not row or not row[0]:
                continue
            try:
                text = str(row[0]).strip()
                opt_a = str(row[1] or '').strip()
                opt_b = str(row[2] or '').strip()
                opt_c = str(row[3] or '').strip()
                opt_d = str(row[4] or '').strip()
                correct = str(row[5] or '').strip().upper()
                explanation = str(row[6] or '').strip()
                difficulty = str(row[7] or 'medium').strip().lower()
                if difficulty not in ('easy', 'medium', 'hard'):
                    difficulty = 'medium'
                if correct not in ('A', 'B', 'C', 'D'):
                    errors.append(f"Qator {row_idx}: to'g'ri javob A/B/C/D bo'lishi kerak")
                    continue
                Question.objects.create(
                    test=test,
                    question_type='multiple_choice',
                    text=text,
                    option_a=opt_a, option_b=opt_b, option_c=opt_c, option_d=opt_d,
                    correct_option=correct,
                    explanation=explanation,
                    difficulty=difficulty,
                    order=test.questions.count() + 1,
                )
                created += 1
            except Exception as e:
                errors.append(f"Qator {row_idx}: {e}")

        return Response({'created': created, 'errors': errors})


class ExcelTemplateView(APIView):
    """GET /api/teacher/excel-template/ — Download blank question import template."""
    permission_classes = [IsTeacher]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Savollar"
        headers = ['Savol matni *', 'A variant *', 'B variant *', 'C variant', 'D variant',
                   'To\'g\'ri javob (A/B/C/D) *', 'Tushuntirish', 'Qiyinlik (easy/medium/hard)']
        hfill = PatternFill("solid", fgColor="4F46E5")
        hfont = Font(color="FFFFFF", bold=True)
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = hfill; c.font = hfont
            c.alignment = Alignment(horizontal='center')
            ws.column_dimensions[c.column_letter].width = max(len(h) + 4, 18)
        # Example row
        ws.append(['2 + 2 = ?', '3', '4', '5', '6', 'B', 'To\'g\'ri javob 4, ya\'ni B', 'easy'])
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="savollar_shablon.xlsx"'
        wb.save(response)
        return response


# ═══════════════════════════════════════════════════════════════════════════════
# AI CHATBOT VIEWS
# ═══════════════════════════════════════════════════════════════════════════════

class ChatView(APIView):
    """POST /api/chat/<attempt_id>/ — Send a message to AI chatbot."""
    permission_classes = [IsStudent]

    def post(self, request, attempt_id):
        from django.conf import settings

        attempt = Attempt.objects.filter(pk=attempt_id, student=request.user).first()
        if not attempt:
            return Response({'detail': 'Urinish topilmadi.'}, status=404)

        chatbot_mode = attempt.test.chatbot_mode
        if chatbot_mode == 'OFF':
            return Response({'detail': 'Chatbot bu test uchun o\'chirilgan.'}, status=403)

        user_message = request.data.get('message', '').strip()
        if not user_message:
            return Response({'detail': 'Xabar bo\'sh bo\'lmasligi kerak.'}, status=400)

        # Get or create chat session
        session, _ = ChatSession.objects.get_or_create(attempt=attempt)

        # Extract context
        current_q_id = request.data.get('current_question_id')
        context_q = None

        # Check if user mentions a specific number (e.g. "10-savol")
        import re
        match = re.search(r'(\d+)-savol', user_message.lower())
        if match:
            idx = int(match.group(1)) - 1
            if 0 <= idx < len(attempt.question_order):
                q_id = attempt.question_order[idx]
                context_q = Question.objects.filter(id=q_id).first()

        if not context_q and current_q_id:
            context_q = Question.objects.filter(id=current_q_id).first()

        q_context_text = ""
        if context_q:
            q_context_text = (
                f"\n\nKONTEKST (O'quvchi so'rayotgan savol):\n"
                f"Savol: {context_q.text}\n"
                f"Variantlar: A: {context_q.option_a}, B: {context_q.option_b}, C: {context_q.option_c}, D: {context_q.option_d}\n"
                f"To'g'ri javob: {context_q.correct_option}\n"
                f"Tushuntirish: {context_q.explanation}\n"
            )

        # Save user message
        ChatMessage.objects.create(session=session, role='user', content=user_message)

        # Build system prompt based on chatbot_mode
        if chatbot_mode == 'HINT_ONLY':
            system_prompt = (
                "Siz Studex ta'lim platformasi yordamchisisiz. "
                "Foydalanuvchiga faqat maslahat, yo'nalish yoki qoida bering. "
                "Hech qachon to'g'ri javobni to'g'ridan-to'g'ri aytmang!"
            )
        else:
            system_prompt = (
                "Siz Studex ta'lim platformasi yordamchisisiz. "
                "Foydalanuvchiga qadam-baqadam to'liq tushuntiring va o'rgating. "
                "yechimni ko'rsating. "
                "O'xshash mashqlar taklif qiling. "
                "O'zbek tilida javob bering."
            )

        # Build conversation history as plain text for Gemini
        messages = session.messages.order_by('created_at')[:20]  # Last 20 messages
        history_text = ""
        for msg in messages:
            role_label = "O'quvchi" if msg.role == 'user' else "Yordamchi"
            history_text += f"{role_label}: {msg.content}\n"

        # Full prompt: context + system + history + current message
        full_prompt = (
            f"{system_prompt}"
            f"{q_context_text}\n\n"
            f"TARIX:\n{history_text}"
            f"O'quvchi: {user_message}\n"
            f"Yordamchi:"
        )

        # Call Google Gemini API
        api_key = getattr(settings, 'GEMINI_API_KEY', '') or os.environ.get('GEMINI_API_KEY', '')
        if not api_key:
            # Fallback response when no API key
            reply = (
                "Kechirasiz, AI xizmati hozir mavjud emas. "
                "O'qituvchingizga murojaat qiling yoki darslikdan foydalaning."
            )
        else:
            import time
            from google import genai as _genai

            # Google GenAI SDK orqali chaqiruv
            # Asosiy model, keyin zaxira
            models_to_try = ['gemini-2.0-flash', 'gemini-2.0-flash-lite', 'gemini-1.5-flash', 'gemini-3-flash-preview' ]
            reply = None
            last_error_code = None

            client = _genai.Client(api_key=api_key)

            for model_name in models_to_try:
                if reply is not None:
                    break
                for attempt_num in range(2):
                    try:
                        response = client.models.generate_content(
                            model=model_name,
                            contents=full_prompt,
                        )
                        reply = response.text
                        break  # Muvaffaqiyatli
                    except Exception as e:
                        err = str(e)
                        if '429' in err or 'RESOURCE_EXHAUSTED' in err:
                            last_error_code = 429
                            break  # Keyingi modelga o'tish
                        elif '503' in err or 'UNAVAILABLE' in err or '500' in err:
                            last_error_code = 503
                            if attempt_num == 0:
                                time.sleep(2)  # Bir oz kut va qayta urin
                                continue
                            break
                        else:
                            last_error_code = 0
                            break

            if reply is None:
                if last_error_code == 429:
                    reply = (
                        "AI xizmati bugunlik limitga yetdi. "
                        "Iltimos, bir soatdan so'ng qayta urinib ko'ring."
                    )
                elif last_error_code == 503:
                    reply = (
                        "AI yordamchisi hozir juda band. "
                        "Bir necha daqiqadan so'ng qayta urinib ko'ring."
                    )
                else:
                    reply = "AI xizmati vaqtincha ishlamayapti. Keyinroq urinib ko'ring."


        # Save assistant message
        ChatMessage.objects.create(session=session, role='assistant', content=reply)

        return Response({'reply': reply})


# ═══════════════════════════════════════════════════════════════════════════════
# ADVANCED ANALYTICS VIEWS
# ═══════════════════════════════════════════════════════════════════════════════

class TeacherAnalyticsView(APIView):
    """GET /api/teacher/analytics/ — Advanced teacher analytics by topic."""
    permission_classes = [IsTeacher]

    def get(self, request):
        teacher = request.user
        # Eng ko'p xato ishlangan mavzular
        wrong_topics = Answer.objects.filter(
            attempt__test__teacher=teacher,
            is_correct=False,
            question__topic__isnull=False
        ).exclude(question__topic="").values('question__topic').annotate(wrong_count=Count('id')).order_by('-wrong_count')
        
        return Response({
            'wrong_topics_chart': list(wrong_topics)
        })


class StudentAnalyticsView(APIView):
    """GET /api/student/analytics/ — Student weak topics and suggested materials."""
    permission_classes = [IsStudent]

    def get(self, request):
        student = request.user
        # Find weakest topics
        wrong_answers = Answer.objects.filter(
            attempt__student=student,
            is_correct=False,
            question__topic__isnull=False
        ).exclude(question__topic="").values('question__topic').annotate(wrong_count=Count('id')).order_by('-wrong_count')[:5]

        weak_topics = [item['question__topic'] for item in wrong_answers]
        
        # Suggest materials based on weakest topics
        suggested = []
        if weak_topics:
            query = Q()
            for t in weak_topics:
                # Oddiy LIKE qidiruv, qisman moslik uchun
                query |= Q(title__icontains=t) | Q(description__icontains=t)
            
            # Faqat shu talabaga mos keluvchi materiallar
            qs = Material.objects.filter(
                Q(class_name=student.class_name) | Q(class_name='') | Q(classroom__in=student.classrooms.all())
            ).filter(query).distinct()[:5]
            
            suggested = MaterialSerializer(qs, many=True).data

        return Response({
            'weak_topics': list(wrong_answers),
            'suggested_materials': suggested
        })


class PublicTestQuestionsView(APIView):
    """GET /api/student/tests/<id>/public-questions/ — Returns questions for public tests."""
    permission_classes = [permissions.AllowAny]

    def get(self, request, test_id):
        test = Test.objects.filter(pk=test_id, is_active=True, access_type='public').first()
        if not test:
            return Response({'detail': 'Test topilmadi yoki bu sinf uchun yopilgan.'}, status=404)

        questions = Question.objects.filter(test=test).prefetch_related('matching_pairs')
        # Return randomized order for guest too
        q_list = list(questions)
        import random
        random.shuffle(q_list)

        return Response({
            'test_name': test.name,
            'time_limit': test.time_limit,
            'chatbot_mode': test.chatbot_mode,
            'questions': QuestionSerializer(q_list, many=True).data,
        })
